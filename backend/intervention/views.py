from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from .models import FicheIntervention
from intervention.serializers import (
    FicheInterventionCreateSerializer,
    FicheInterventionDetailSerializer,
    FicheInterventionUpdateSerializer
)
from django.contrib.auth import get_user_model
from users.serializers import UserSerializer
from users.permissions import IsAdminOrInstallateur,IsInstallateur
from rest_framework.permissions import IsAuthenticated
from rest_framework import generics, permissions
from dateutil.relativedelta import relativedelta  
from datetime import datetime

User = get_user_model()

class ListeFicheInterventionView(generics.ListAPIView):
    """Liste toutes les fiches d'intervention"""
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = FicheIntervention.objects.all()
        
        search_term = self.request.query_params.get('search')
        
        if search_term:
            queryset = queryset.filter(
                Q(installation__nom__icontains=search_term) |
                Q(statut__icontains=search_term) |
                Q(technicien__email__icontains=search_term) |
                Q(type_intervention__icontains=search_term) |
                Q(technicien__first_name__icontains=search_term) |
                Q(technicien__last_name__icontains=search_term)
            )
            
        technicien_id = self.request.query_params.get('technicien')
        if technicien_id:
            queryset = queryset.filter(technicien_id=technicien_id)
            
        installation_id = self.request.query_params.get('installation')
        if installation_id:
            queryset = queryset.filter(installation_id=installation_id)
            
        statut = self.request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        type_intervention = self.request.query_params.get('type_intervention')
        if type_intervention:
            queryset = queryset.filter(type_intervention=type_intervention)

            
        return queryset

from .utils import notifier_creation_intervention

class CreerFicheInterventionView(generics.CreateAPIView):
    serializer_class = FicheInterventionCreateSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrInstallateur]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        instance = serializer.instance

        notifier_creation_intervention(instance)

        detail_serializer = FicheInterventionDetailSerializer(instance)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from .models import FicheIntervention
from .serializers import FicheInterventionDetailSerializer

class DetailFicheInterventionView(generics.RetrieveAPIView):
    """Récupérer les détails d'une fiche d'intervention avec ses statistiques"""
    queryset = FicheIntervention.objects.all()
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        installation = instance.installation

        interventions_passees = FicheIntervention.objects.filter(
            installation=installation
        ).exclude(id=instance.id)

        stats = {
            "nombre_interventions_precedentes": interventions_passees.count(),
            "temps_depuis_derniere": None,
            "temps_moyen_entre": None,
        }

        derniere = interventions_passees.order_by('-date_prevue').first()
        if derniere:
            stats["temps_depuis_derniere"] = (instance.date_prevue - derniere.date_prevue).days

        if interventions_passees.count() > 0:
            dates = list(interventions_passees.values_list('date_prevue', flat=True).order_by('date_prevue'))
            dates.append(instance.date_prevue)
            deltas = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
            stats["temps_moyen_entre"] = sum(deltas) // len(deltas) if deltas else 0

        data = self.get_serializer(instance).data
        data["statistiques"] = stats
        return Response(data)


class ModifierFicheInterventionView(generics.UpdateAPIView):
    """Modifier une fiche d'intervention existante"""
    queryset = FicheIntervention.objects.all()
    serializer_class = FicheInterventionUpdateSerializer
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        instance = self.get_object()
        detail_serializer = FicheInterventionDetailSerializer(instance)
        return Response(detail_serializer.data)

class SupprimerFicheInterventionView(generics.DestroyAPIView):
    """Supprimer une fiche d'intervention"""
    queryset = FicheIntervention.objects.all()
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

class ChangerStatutFicheInterventionView(APIView):
    """Changer le statut d'une fiche d'intervention"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        fiche = get_object_or_404(FicheIntervention, pk=pk)
        nouveau_statut = request.data.get('statut')
        
        if nouveau_statut not in dict(FicheIntervention.STATUT_CHOICES):
            return Response(
                {'error': 'Statut invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        fiche.statut = nouveau_statut
        fiche.save()
        
        serializer = FicheInterventionDetailSerializer(fiche) 
        return Response(serializer.data)


class TechniciensListView(APIView):
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def get(self, request):
        techniciens = User.objects.filter(groups__name='Techniciens')
        serializer = UserSerializer(techniciens, many=True)
        return Response({"results": serializer.data})


class HistoriqueInterventionsParInstallationView(generics.ListAPIView):
    """Afficher l'historique des interventions pour une installation donnée"""
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def get_queryset(self):
        installation_id = self.kwargs.get('installation_id')
        return FicheIntervention.objects.filter(installation_id=installation_id).order_by('-date_prevue')


from django.db.models import Count, Q

class NombreInterventionsParTechnicienView(APIView):
    """Retourne le nombre total d'interventions par technicien"""
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def get(self, request):
        data = (
            FicheIntervention.objects.values('technicien__id', 'technicien__first_name', 'technicien__last_name')
            .annotate(nombre_interventions=Count('id'))
        )

        result = [
            {
                "technicien_id": entry["technicien__id"],
                "nom": f'{entry["technicien__first_name"]} {entry["technicien__last_name"]}',
                "nombre_interventions": entry["nombre_interventions"]
            }
            for entry in data
        ]

        return Response(result)


class TauxResolutionInterventionsView(APIView):
    """Retourne le taux de résolution des interventions"""
    permission_classes = [permissions.IsAuthenticated,IsAdminOrInstallateur]

    def get(self, request):
        total = FicheIntervention.objects.count()
        resolues = FicheIntervention.objects.filter(statut='resolue').count()

        taux_resolution = (resolues / total * 100) if total > 0 else 0

        return Response({
            "total_interventions": total,
            "interventions_resolues": resolues,
            "taux_resolution": f"{taux_resolution:.2f} %"
        })
    
class ListeFicheInterventionClientView(generics.ListAPIView):
    """
    Retourne les fiches d'intervention liées aux installations du client connecté
    """
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return FicheIntervention.objects.filter(installation__client=user).order_by('-date_prevue')
    
class DetailFicheInterventionClientView(generics.RetrieveAPIView):
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return FicheIntervention.objects.filter(installation__client=user)


#liste intervention installateur 

class ListeMesFichesInterventionView(generics.ListAPIView):
    """
    Liste des fiches d'intervention pour les installations de l'installateur connecté.
    """
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [IsAuthenticated, IsInstallateur]

    def get_queryset(self):
        user = self.request.user

        if user.role != 'installateur':
            return FicheIntervention.objects.none()

        return FicheIntervention.objects.filter(installation__installateur=user).order_by('-date_prevue')


from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import openpyxl
import csv
from openpyxl.utils import get_column_letter
from django.db.models import Q
from .models import FicheIntervention
# Export CSV
class ExportInterventionsCSVView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = self._filter_queryset(request)
        return self._export_csv(queryset)

    def _filter_queryset(self, request):
        qs = FicheIntervention.objects.all()
        statut = request.query_params.get('statut')
        technicien = request.query_params.get('technicien')
        type_intervention = request.query_params.get('type_intervention')
        search = request.query_params.get('search')

        if statut:
            qs = qs.filter(statut=statut)
        if technicien:
            qs = qs.filter(technicien_id=technicien)
        if type_intervention:
            qs = qs.filter(type_intervention=type_intervention)
        if search:
            qs = qs.filter(
                Q(installation__nom__icontains=search) |
                Q(technicien__first_name__icontains=search) |
                Q(technicien__last_name__icontains=search) |
                Q(description__icontains=search)
            )
        return qs

    def _export_csv(self, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="interventions.csv"'
        writer = csv.writer(response)
        writer.writerow(["Installation", "Technicien", "Date prévue", "Type", "Statut", "Description"])
        for obj in queryset:
            writer.writerow([
                obj.installation.nom if obj.installation else "",
                f"{obj.technicien.first_name} {obj.technicien.last_name}" if obj.technicien else "",
                obj.date_prevue.strftime("%Y-%m-%d %H:%M") if obj.date_prevue else "",
                obj.get_type_intervention_display(),
                obj.get_statut_display(),
                obj.description or ""
            ])
        return response


# Export XLSX
class ExportInterventionsXLSXView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = ExportInterventionsCSVView()._filter_queryset(request)  # Réutilise le filtre
        return self._export_xlsx(queryset)

    def _export_xlsx(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interventions"
        headers = ["Installation", "Technicien", "Date prévue", "Type", "Statut", "Description"]
        ws.append(headers)

        for obj in queryset:
            ws.append([
                obj.installation.nom if obj.installation else "",
                f"{obj.technicien.first_name} {obj.technicien.last_name}" if obj.technicien else "",
                obj.date_prevue.strftime("%Y-%m-%d %H:%M") if obj.date_prevue else "",
                obj.get_type_intervention_display(),
                obj.get_statut_display(),
                obj.description or ""
            ])

        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="interventions.xlsx"'
        return response
    

# Technicien

from users.permissions import IsTechnicien
class ListeFichesInterventionTechnicienView(generics.ListAPIView):
    """
    Liste des fiches d'intervention du technicien connecté.
    """
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [IsAuthenticated, IsTechnicien]

    def get_queryset(self):
        user = self.request.user

        return FicheIntervention.objects.filter(technicien=user).order_by('-date_prevue')
    
class DetailFicheInterventionTechnicienView(generics.RetrieveAPIView):
    serializer_class = FicheInterventionDetailSerializer
    permission_classes = [IsAuthenticated,IsTechnicien]

    def get_queryset(self):
        user = self.request.user
        return FicheIntervention.objects.filter(technicien=user)
