from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum
from django.db.models.functions import TruncDay
from production.models import ProductionConsommation
from users.permissions import IsAdminOrInstallateur
from datetime import datetime
import pandas as pd
import io
from django.http import HttpResponse
 
class RapportProductionMensuelleView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def get(self, request):
        installation_id = request.query_params.get("installation_id")
        mois = request.query_params.get("mois")  # Format: "2025-03"
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois sont requis"}, status=400)
 
        try:
            year, month = map(int, mois.split("-"))
 
            queryset = ProductionConsommation.objects.filter(
                installation_id=installation_id,
                horodatage__year=year,
                horodatage__month=month
            ).annotate(day=TruncDay("horodatage")).values("day").annotate(
                total=Sum("energie_produite_kwh")
            ).order_by("day")
 
            resultats = [
                {
                    "jour": entry["day"].strftime("%Y-%m-%d"),
                    "production_kwh": float(entry["total"])
                }
                for entry in queryset
            ]
 
            return Response(resultats, status=200)
 
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
class ExporterRapportProductionView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def post(self, request):
        installation_id = request.data.get("installation_id")
        mois = request.data.get("mois")  # ex: "2025-03"
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=400)
 
        year, month = map(int, mois.split("-"))
 
        queryset = ProductionConsommation.objects.filter(
            installation_id=installation_id,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            production_kwh=Sum("energie_produite_kwh")
        ).order_by("jour")
 
        data = [{
            "Jour": entry["jour"].strftime("%Y-%m-%d"),
            "Production (kWh)": float(entry["production_kwh"])
        } for entry in queryset]
 
        df = pd.DataFrame(data)
 
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Production", index=False)
 
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"rapport_production_{mois}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.http import HttpResponse
import tempfile
from rest_framework.views import APIView

from django.db.models.functions import TruncDay
from django.db.models import Sum
from rest_framework.response import Response
 
class ExporterRapportProductionPDFView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def post(self, request):
        installation_id = request.data.get("installation_id")
        mois = request.data.get("mois")  # format YYYY-MM
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=400)
 
        year, month = map(int, mois.split("-"))
        queryset = ProductionConsommation.objects.filter(
            installation_id=installation_id,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            production_kwh=Sum("energie_produite_kwh")
        ).order_by("jour")
 
        data = [{
            "jour": entry["jour"].strftime("%Y-%m-%d"),
            "production_kwh": float(entry["production_kwh"])
        } for entry in queryset]
 
        html_string = render_to_string("rapport_production_pdf.html", {
            "mois": mois,
            "donnees": data,
        })
 
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_production_{mois}.pdf"'
 
        pisa_status = pisa.CreatePDF(
            html_string, dest=response
        )
 
        if pisa_status.err:
            return Response({'error': 'Erreur lors de la génération du PDF'}, status=500)
        return response
    
class RapportConsommationMensuelleView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def get(self, request):
        installation_id = request.query_params.get("installation_id")
        mois = request.query_params.get("mois")  # ex: 2025-03
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=400)
 
        year, month = map(int, mois.split("-"))
        queryset = ProductionConsommation.objects.filter(
            installation_id=installation_id,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            consommation_kwh=Sum("energie_consomme_kwh")
        ).order_by("jour")
 
        resultats = [
            {
                "jour": entry["jour"].strftime("%Y-%m-%d"),
                "consommation_kwh": float(entry["consommation_kwh"])
            }
            for entry in queryset
        ]
 
        return Response(resultats, status=200)
    
 
class ExporterRapportConsommationExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def post(self, request):
        installation_id = request.data.get("installation_id")
        mois = request.data.get("mois")
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=400)
 
        year, month = map(int, mois.split("-"))
        queryset = ProductionConsommation.objects.filter(
            installation_id=installation_id,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            consommation_kwh=Sum("energie_consomme_kwh")
        ).order_by("jour")
 
        data = [{
            "Date": entry["jour"].strftime("%Y-%m-%d"),
            "Consommation (kWh)": float(entry["consommation_kwh"])
        } for entry in queryset]
 
        df = pd.DataFrame(data)
 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=rapport_consommation_{mois}.xlsx'
        df.to_excel(response, index=False)
 
        return response
    
class ExporterRapportConsommationPDFView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrInstallateur]
 
    def post(self, request):
        installation_id = request.data.get("installation_id")
        mois = request.data.get("mois")
 
        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=400)
 
        year, month = map(int, mois.split("-"))
        queryset = ProductionConsommation.objects.filter(
            installation_id=installation_id,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            consommation_kwh=Sum("energie_consomme_kwh")
        ).order_by("jour")
 
        data = [{
            "jour": entry["jour"].strftime("%Y-%m-%d"),
            "consommation_kwh": float(entry["consommation_kwh"])
        } for entry in queryset]
 
        html_string = render_to_string("rapport_consommation_pdf.html", {
            "mois": mois,
            "donnees": data,
        })
 
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=rapport_consommation_{mois}.pdf'
 
        pisa_status = pisa.CreatePDF(html_string, dest=response)
 
        if pisa_status.err:
            return Response({'error': 'Erreur PDF'}, status=500)
        return response
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from django.utils.timezone import make_aware
from django.db.models import F
from alarme.models import AlarmeDeclenchee
import calendar

class RapportAlarmesMensuellesView(APIView):
    def get(self, request):
        installation_id = request.query_params.get('installation_id')
        mois = request.query_params.get('mois')  # Format "YYYY-MM"

        if not installation_id or not mois:
            return Response({"error": "installation_id et mois requis"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            annee, mois_int = map(int, mois.split("-"))
            date_debut = make_aware(datetime(annee, mois_int, 1))
            dernier_jour = calendar.monthrange(annee, mois_int)[1]
            date_fin = make_aware(datetime(annee, mois_int, dernier_jour, 23, 59, 59))
        except:
            return Response({"error": "Format de date invalide"}, status=status.HTTP_400_BAD_REQUEST)

        alarmes = (
            AlarmeDeclenchee.objects
            .filter(installation_id=installation_id, date_declenchement__range=(date_debut, date_fin))
            .select_related("code_alarme")
            .order_by("date_declenchement")
        )

        donnees = [
            {
                "date": alarme.date_declenchement.strftime("%Y-%m-%d %H:%M"),
                "gravite": alarme.code_alarme.gravite,
                "type": alarme.code_alarme.type_alarme,
                "statut": "Résolue" if alarme.est_resolue else "Non résolue"
            }
            for alarme in alarmes
        ]

        return Response(donnees, status=status.HTTP_200_OK)



class ExportRapportAlarmesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            installation_id = request.data.get("installation_id")
            mois = request.data.get("mois")

            if not installation_id or not mois:
                return Response({"error": "Champs 'installation_id' et 'mois' requis."}, status=400)

            try:
                annee, mois_num = map(int, mois.split("-"))
                date_debut = make_aware(datetime(annee, mois_num, 1))
                dernier_jour = calendar.monthrange(annee, mois_num)[1]
                date_fin = make_aware(datetime(annee, mois_num, dernier_jour, 23, 59, 59))
            except ValueError:
                return Response({"error": "Format de date invalide. Utiliser 'YYYY-MM'."}, status=400)

            alarmes = (
                AlarmeDeclenchee.objects
                .filter(installation_id=installation_id, date_declenchement__range=(date_debut, date_fin))
                .select_related("code_alarme")
            )

            if not alarmes.exists():
                return Response({"error": "Aucune alarme trouvée pour cette période."}, status=404)

            data = []
            for alarme in alarmes:
                data.append({
                    "Date": alarme.date_declenchement.strftime("%Y-%m-%d %H:%M") if alarme.date_declenchement else "",
                    "Gravité": alarme.code_alarme.gravite,
                    "Type": alarme.code_alarme.type_alarme,
                    "Statut": "Résolue" if alarme.est_resolue else "Non résolue",
                    "Description": getattr(alarme.code_alarme, "description", "") or "",
                })

            df = pd.DataFrame(data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name="Alarmes")

            output.seek(0)
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="rapport_alarmes_{mois}.xlsx"'
            return response

        except Exception as e:
            print("❌ Erreur export Excel :", str(e))
            return Response({"error": f"Erreur serveur : {str(e)}"}, status=500)
    
from reportlab.pdfgen import canvas

class ExportRapportAlarmesPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        installation_id = request.data.get("installation_id")
        mois = request.data.get("mois")

        if not installation_id or not mois:
            return HttpResponse("Installation et mois requis", status=400)

        annee, mois_num = map(int, mois.split("-"))
        debut = make_aware(datetime(annee, mois_num, 1))
        fin = make_aware(datetime(annee, mois_num, calendar.monthrange(annee, mois_num)[1], 23, 59, 59))

        alarmes = (
            AlarmeDeclenchee.objects
            .filter(installation_id=installation_id, date_declenchement__range=(debut, fin))
            .select_related("code_alarme")
        )

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)
        p.setTitle(f"Rapport Alarmes - {mois}")

        p.drawString(100, 800, f"Rapport des alarmes - Mois {mois}")

        y = 780
        for a in alarmes:
            line = f"{a.date_declenchement.strftime('%Y-%m-%d %H:%M')} - {a.code_alarme.gravite} - {a.code_alarme.type_alarme} - {'Résolue' if a.est_resolue else 'Non résolue'}"
            p.drawString(50, y, line)
            y -= 20
            if y < 50:
                p.showPage()
                y = 800

        p.save()
        buffer.seek(0)

        return HttpResponse(buffer, content_type='application/pdf', headers={
            'Content-Disposition': f'attachment; filename="rapport_alarmes_{mois}.pdf"'
        })

#Client

from installations.models import Installation

class RapportProductionClientView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        mois = request.query_params.get("mois") 

        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        try:
            year, month = map(int, mois.split("-"))
            installation = Installation.objects.filter(client=request.user).first()
            if not installation:
                return Response({"error": "Aucune installation trouvée pour ce client."}, status=404)

            queryset = ProductionConsommation.objects.filter(
                installation=installation,
                horodatage__year=year,
                horodatage__month=month
            ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
                production_kwh=Sum("energie_produite_kwh")
            ).order_by("jour")

            resultats = [
                {
                    "jour": entry["jour"].strftime("%Y-%m-%d"),
                    "production_kwh": float(entry["production_kwh"])
                }
                for entry in queryset
            ]

            return Response(resultats, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
class RapportConsommationClientView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        mois = request.query_params.get("mois")

        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        try:
            year, month = map(int, mois.split("-"))
            installation = Installation.objects.filter(client=request.user).first()
            if not installation:
                return Response({"error": "Aucune installation trouvée pour ce client."}, status=404)

            queryset = ProductionConsommation.objects.filter(
                installation=installation,
                horodatage__year=year,
                horodatage__month=month
            ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
                consommation_kwh=Sum("energie_consomme_kwh")
            ).order_by("jour")

            resultats = [
                {
                    "jour": entry["jour"].strftime("%Y-%m-%d"),
                    "consommation_kwh": float(entry["consommation_kwh"])
                }
                for entry in queryset
            ]

            return Response(resultats, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
class RapportAlarmesClientView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        mois = request.query_params.get("mois")  # Format "YYYY-MM"
        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        try:
            year, month = map(int, mois.split("-"))
            date_debut = make_aware(datetime(year, month, 1))
            date_fin = make_aware(datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59))

            installation = Installation.objects.filter(client=request.user).first()
            if not installation:
                return Response({"error": "Aucune installation trouvée pour ce client."}, status=404)

            alarmes = AlarmeDeclenchee.objects.filter(
                installation=installation,
                date_declenchement__range=(date_debut, date_fin)
            ).select_related("code_alarme")

            donnees = [{
                "date": a.date_declenchement.strftime("%Y-%m-%d %H:%M"),
                "gravite": a.code_alarme.gravite,
                "type": a.code_alarme.type_alarme,
                "statut": "Résolue" if a.est_resolue else "Non résolue"
            } for a in alarmes]

            return Response(donnees, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)



class ExportRapportProductionClientExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        year, month = map(int, mois.split("-"))
        installation = Installation.objects.filter(client=request.user).first()
        if not installation:
            return Response({"error": "Installation non trouvée."}, status=404)

        queryset = ProductionConsommation.objects.filter(
            installation=installation,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            production_kwh=Sum("energie_produite_kwh")
        ).order_by("jour")

        data = [{"Date": e["jour"].strftime("%Y-%m-%d"), "Production (kWh)": float(e["production_kwh"])} for e in queryset]
        df = pd.DataFrame(data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=rapport_production_{mois}.xlsx'
        df.to_excel(response, index=False)
        return response

class ExportRapportConsommationClientExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        year, month = map(int, mois.split("-"))
        installation = Installation.objects.filter(client=request.user).first()
        if not installation:
            return Response({"error": "Installation non trouvée."}, status=404)

        queryset = ProductionConsommation.objects.filter(
            installation=installation,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            consommation_kwh=Sum("energie_consomme_kwh")
        ).order_by("jour")

        data = [{"Date": e["jour"].strftime("%Y-%m-%d"), "Consommation (kWh)": float(e["consommation_kwh"])} for e in queryset]
        df = pd.DataFrame(data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=rapport_consommation_{mois}.xlsx'
        df.to_excel(response, index=False)
        return response

class ExportRapportAlarmesClientExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        year, month = map(int, mois.split("-"))
        debut = make_aware(datetime(year, month, 1))
        fin = make_aware(datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59))

        installation = Installation.objects.filter(client=request.user).first()
        if not installation:
            return Response({"error": "Installation non trouvée."}, status=404)

        alarmes = AlarmeDeclenchee.objects.filter(
            installation=installation,
            date_declenchement__range=(debut, fin)
        ).select_related("code_alarme")

        data = []
        for a in alarmes:
            data.append({
                "Date": a.date_declenchement.strftime("%Y-%m-%d %H:%M"),
                "Gravité": a.code_alarme.gravite,
                "Type": a.code_alarme.type_alarme,
                "Statut": "Résolue" if a.est_resolue else "Non résolue"
            })

        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name="Alarmes")

        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=rapport_alarmes_{mois}.xlsx'
        return response
    
class ExportRapportProductionClientPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "Le paramètre 'mois' est requis."}, status=400)

        try:
            year, month = map(int, mois.split("-"))
            installation = Installation.objects.filter(client=request.user).first()
            if not installation:
                return Response({"error": "Installation non trouvée."}, status=404)

            queryset = ProductionConsommation.objects.filter(
                installation=installation,
                horodatage__year=year,
                horodatage__month=month
            ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
                production_kwh=Sum("energie_produite_kwh")
            ).order_by("jour")

            data = [{"jour": e["jour"].strftime("%Y-%m-%d"), "production_kwh": float(e["production_kwh"])} for e in queryset]

            html = render_to_string("rapport_production_client.html", {
                "mois": mois,
                "donnees": data,
                "installation": installation
            })

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="rapport_production_{mois}.pdf"'

            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return Response({"error": "Erreur génération PDF"}, status=500)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)

class ExportRapportConsommationClientPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "mois est requis"}, status=400)

        year, month = map(int, mois.split("-"))
        installation = Installation.objects.filter(client=request.user).first()
        if not installation:
            return Response({"error": "Installation non trouvée"}, status=404)

        queryset = ProductionConsommation.objects.filter(
            installation=installation,
            horodatage__year=year,
            horodatage__month=month
        ).annotate(jour=TruncDay("horodatage")).values("jour").annotate(
            consommation_kwh=Sum("energie_consomme_kwh")
        ).order_by("jour")

        data = [{"jour": e["jour"].strftime("%Y-%m-%d"), "consommation_kwh": float(e["consommation_kwh"])} for e in queryset]

        html = render_to_string("rapport_consommation_client.html", {
            "mois": mois,
            "donnees": data,
            "installation": installation
        })

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="rapport_consommation_{mois}.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return Response({"error": "Erreur génération PDF"}, status=500)
        return response

class ExportRapportAlarmesClientPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        mois = request.data.get("mois")
        if not mois:
            return Response({"error": "mois est requis"}, status=400)

        year, month = map(int, mois.split("-"))
        debut = make_aware(datetime(year, month, 1))
        fin = make_aware(datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59))

        installation = Installation.objects.filter(client=request.user).first()
        if not installation:
            return Response({"error": "Installation non trouvée"}, status=404)

        alarmes = AlarmeDeclenchee.objects.filter(
            installation=installation,
            date_declenchement__range=(debut, fin)
        ).select_related("code_alarme")

        data = [{
            "date": a.date_declenchement.strftime("%Y-%m-%d %H:%M"),
            "gravite": a.code_alarme.gravite,
            "type": a.code_alarme.type_alarme,
            "statut": "Résolue" if a.est_resolue else "Non résolue"
        } for a in alarmes]

        html = render_to_string("rapport_alarmes_client.html", {
            "mois": mois,
            "donnees": data,
            "installation": installation
        })

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="rapport_alarmes_{mois}.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return Response({"error": "Erreur génération PDF"}, status=500)
        return response


#technicien

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from intervention.models import FicheIntervention
from entretien.models import Entretien
from datetime import datetime
from django.db.models import Count, Avg
from users.permissions import IsTechnicien
from django.db.models.functions import TruncMonth

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rapport_technique_technicien(request):
    user = request.user

    if user.role != 'technicien':
        return Response({"error": "Accès réservé aux techniciens"}, status=403)

    interventions = FicheIntervention.objects.filter(technicien=user)
    entretiens = Entretien.objects.filter(technicien=user)

    # KPI
    kpi = {
        "total_interventions": interventions.count(),
        "total_entretiens": entretiens.count(),
        "moyenne_duree_entretiens": entretiens.aggregate(avg=Avg("duree_estimee"))["avg"] or 0
    }

    # Groupement par type
    types_intervention = interventions.values("type_intervention").annotate(count=Count("id"))
    types_entretien = entretiens.values("type_entretien").annotate(count=Count("id"))

    # Groupement par mois (PostgreSQL compatible)
    interventions_par_mois = (
        interventions
        .annotate(month=TruncMonth("date_prevue"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    entretiens_par_mois = (
        entretiens
        .annotate(month=TruncMonth("date_debut"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    def format_month_data(queryset):
        return [
            {
                "month": item["month"].strftime("%Y-%m"),
                "count": item["count"]
            } for item in queryset
        ]

    return Response({
        "kpi": kpi,
        "graphiques": {
            "types_intervention": list(types_intervention),
            "types_entretien": list(types_entretien),
            "interventions_par_mois": format_month_data(interventions_par_mois),
            "entretiens_par_mois": format_month_data(entretiens_par_mois),
        }
    })


import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

@api_view(["GET"])
@permission_classes([IsAuthenticated, IsTechnicien])
def export_rapport_technicien_excel(request):
    user = request.user
    interventions = FicheIntervention.objects.filter(technicien=user)
    entretiens = Entretien.objects.filter(technicien=user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Technique"

    # En-têtes
    headers = [
        "Type", "Installation", "Date prévue", "Statut", "Durée (min)", "Description/Notes"
    ]
    ws.append(headers)

    # Interventions
    for i in interventions:
        ws.append([
            "Intervention",
            i.installation.nom if i.installation else "",
            i.date_prevue.strftime("%Y-%m-%d %H:%M"),
            i.get_statut_display(),
            "",  # Pas de durée estimée sur intervention
            i.description or "",
        ])

    # Entretiens
    for e in entretiens:
        ws.append([
            "Entretien",
            e.installation.nom if e.installation else "",
            e.date_debut.strftime("%Y-%m-%d %H:%M") if e.date_debut else "",
            e.get_statut_display(),
            e.duree_estimee,
            e.notes or "",
        ])

    # Auto-width
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # Réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=rapport_technicien.xlsx"
    wb.save(response)
    return response

from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils.timezone import now
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from entretien.models import Entretien
from intervention.models import FicheIntervention
from xhtml2pdf import pisa 

class ExportRapportTechnicienPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        technicien = request.user

        entretiens = Entretien.objects.filter(technicien=technicien)
        interventions = FicheIntervention.objects.filter(technicien=technicien)

        data_entretiens = [{
            "installation": e.installation.nom,
            "date": e.date_debut.strftime("%Y-%m-%d %H:%M"),
            "type": e.get_type_entretien_display(),
            "statut": e.get_statut_display()
        } for e in entretiens]

        data_interventions = [{
            "installation": i.installation.nom,
            "date": i.date_prevue.strftime("%Y-%m-%d %H:%M"),
            "type": i.get_type_intervention_display(),
            "statut": i.get_statut_display()
        } for i in interventions]

        html = render_to_string("rapport_pdf.html", {
            "technicien": technicien,
            "date": now(),
            "entretiens": data_entretiens,
            "interventions": data_interventions
        })

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="rapport_technicien_{now().date()}.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return Response({"error": "Erreur génération PDF"}, status=500)
        return response
